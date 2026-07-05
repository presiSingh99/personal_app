"""
FORGE — The Daily Ledger (Streamlit edition)
Run locally with:  streamlit run app.py
Every entry auto-saves to a local Excel workbook (default: forge_log.xlsx,
same folder as this script) on every change — no save button needed.
One row per calendar day — updating today's inputs updates today's row.
"""

import os
import datetime as dt

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Forge — Daily Ledger", page_icon="⚒", layout="centered")

DEFAULT_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "forge_log.xlsx")

COLUMNS = [
    "Date", "Weekday", "Workout Split", "Workout Done",
    "Target 1 - Code/Review", "Target 2 - Deep Work", "Target 3 - Water/Nutrition",
    "Gratitude Reflected", "Project Hours", "Study/Review Hours", "Journal",
    "Daily XP %", "Streak",
]

SPLITS = {
    0: ("Active Recovery", "Incline walk 30-45min, mobility, structural rest"),
    1: ("Back Focus", "Lat Pulldown 4x10, Seated Cable Row 4x10, Chest Supported Row 3x10, Single Arm DB Row 3x10, Face Pulls 3x15, Shrugs 3x12"),
    2: ("Chest & Lateral Delts", "Incline DB Press 4x8-10, Flat Press 4x10, Incline Machine Press 3x10, Cable Flys 3x12, Pushups 2xfailure, Lateral Raises 4x15"),
    3: ("Legs Engine", "Squat/Hack Squat 4x8-10, RDL 4x10, Leg Press 3x12, Ham Curl 3x12, Leg Extension 3x12, Calf Raises 4x15, Abs Circuit 10min"),
    4: ("Aesthetic Symmetry", "Incline Press 3x10, Lat Pulldown/Pullups 3x10, Cable Row 3x12, Shoulder Press 3x10, Lateral Raises 5x15, Rear Delt Flys 3x15, Bicep Curl 3x12, Tricep Pushdown 3x12"),
    5: ("Arms & Shoulder Caps", "EZ Bar Curl 4x10, Hammer Curl 3x12, Preacher Curl 3x12, Rope Pushdown 4x12, Overhead Tricep Ext 3x12, Lateral Raises 5x15, Rear Delt Flys 3x15"),
    6: ("Active Recovery", "Incline walk 30-45min, mobility, structural rest"),
}
SPLIT_BY_PY_WEEKDAY = {0: SPLITS[1], 1: SPLITS[2], 2: SPLITS[3], 3: SPLITS[4], 4: SPLITS[5], 5: SPLITS[6], 6: SPLITS[0]}

INTEL = [
    ("Control the input, not the outcome", "The only things fully in your hands each day are your effort, your attention, and your reactions — not results, not other people, not the market. Before you start work, write down one input you fully control today (hours of deep focus, one hard conversation you'll have, one rep you'll add) and judge the day by whether you did that, not by whether the outcome landed. This single reframe removes most anxiety, because anxiety lives in the gap between 'I want X to happen' and 'X is not up to me.' Practice it by ending each day naming the input you controlled well, regardless of the result.",
     "Idempotency",
     "An operation is idempotent if running it once has the exact same effect as running it five times. This matters because networks fail, retries happen, and if a 'charge customer $50' request fires twice due to a timeout, you cannot let it charge $100. The fix is a deterministic key: generate a unique request ID client-side, and have the server check 'have I already processed this ID?' before acting — if yes, return the original result instead of redoing the work. In data pipelines, this looks like using MERGE/UPSERT instead of INSERT, or overwriting a specific partition instead of appending to it. The practical test: if you're afraid to re-run a job, it isn't idempotent yet, and that fear is a signal to fix the design before it fails in production at 3am."),

    ("Do the hard thing first, every day", "Willpower is highest in the morning and decays with every decision you make, so the single highest-leverage habit is doing your most avoided, most important task before you touch email, messages, or anything reactive. This isn't about discipline in the abstract — it's sequencing. Identify the one task you'd most like to postpone, and make a rule that it happens within the first hour of your work day, no exceptions. Within two weeks this compounds: the backlog of avoided things shrinks, and your afternoons become genuinely lighter because you're not carrying dread into them.",
     "Parquet vs CSV",
     "CSV stores data row-by-row as plain text — to read one column, you still have to scan every row and every other column along the way. Parquet stores data column-by-column in a compressed, typed binary format, so a query that only needs 3 of your 40 columns physically reads only those 3, often cutting I/O by 10-100x on large tables. Parquet also stores min/max statistics per column chunk, letting query engines skip entire blocks of data that can't match a filter (predicate pushdown) without even decompressing them. The practical rule: use CSV only as a wire format for ingesting data from external systems, and convert to Parquet the moment it lands in your own storage — never let an analytical query scan raw CSV twice."),

    ("Name the fear specifically", "Vague anxiety ('this project might fail') is paralyzing; specific fear ('if I ship this and the API breaks under load, I'll have to explain it in Monday's review') is solvable, because you can now build a load test. Whenever you notice avoidance or procrastination, stop and write the actual worst-case sentence in concrete terms — who does what happens to whom. Almost always, the specific version has an obvious next action attached to it, while the vague version only has more anxiety attached to it. This is a five-minute practice that turns dread into a task list.",
     "Vector Index Parameters (HNSW)",
     "HNSW (Hierarchical Navigable Small World) builds a multi-layer graph where each vector is connected to its nearest neighbors, letting you search millions of vectors in milliseconds instead of comparing against all of them. The parameter M controls how many neighbor connections each node keeps — higher M means better recall but more memory and slower index builds, and it's fixed once the index is built. The parameter efSearch controls how many candidates are explored at query time — higher efSearch means better recall at the cost of latency, and critically, it can be changed per-query without rebuilding anything. The practical workflow: build your index once with a reasonable M (commonly 16-64), then tune efSearch against your own query distribution and measured recall@k, because synthetic benchmarks rarely match how your users actually query."),

    ("Protect one block of undistracted time daily", "Deep, undistracted focus is a different cognitive mode than reactive, interrupted work, and most people never enter it because notifications never stop. Block 90 minutes on your calendar, physically silence and remove your phone from the room, and tell people in advance you're unreachable in that window. The first week feels uncomfortable because you'll reach for the phantom phone; by week three, that block becomes the most productive 90 minutes of your day, often producing more real output than the other seven hours combined.",
     "RAG Chunking Strategy",
     "Retrieval-Augmented Generation only works as well as the chunks you feed into your vector database — if a chunk cuts a sentence in half or merges two unrelated topics, the embedding for that chunk represents neither topic well, and retrieval quality suffers regardless of how good your model is. Chunk along semantic boundaries (headings, paragraphs, list items) instead of a fixed character count, and keep a 10-20% overlap between consecutive chunks so that ideas spanning a boundary aren't lost entirely to one side. Store the source document, section title, and position as metadata alongside every chunk, since this lets you re-rank, cite sources, and debug bad retrievals later. The uncomfortable truth: most RAG quality problems are chunking problems, and no amount of reranking or bigger models fixes a badly chunked corpus."),

    ("Review your week before planning the next one", "Most people plan forward without ever looking backward, which means the same mistakes repeat silently. Spend 15 minutes every Sunday reviewing what actually happened versus what you planned — what got done, what got skipped, and honestly why. Then plan the next week using that evidence instead of optimism. This one habit, done consistently, is often the single biggest lever for improving how realistic and achievable your plans become over time.",
     "Slowly Changing Dimensions (SCD Type 2)",
     "When a customer's address or a product's price changes, naively overwriting the old value destroys your ability to answer historical questions like 'what did this order actually cost at the time it was placed?' SCD Type 2 solves this by never updating a row in place — instead, it closes the old row (setting a valid_to timestamp) and inserts a brand new row with the new value and a valid_from timestamp. A current_flag column marks which row is active 'now,' and a surrogate key (separate from the natural business key) lets you join fact tables to the exact historical version that was true at that time. This is more storage and more complexity than a simple UPDATE, but it's the only pattern that lets you truthfully reconstruct the past — and analysts will eventually ask a question that requires it."),

    ("Separate the decision from the emotion", "Big decisions made while angry, anxious, or euphoric are usually worse than the same decision made 24 hours later with a clear head, because emotional states distort your estimate of risk and reward in predictable directions. Build a rule: any decision above a certain stakes threshold gets a mandatory 24-hour pause before you act on it, no matter how urgent it feels in the moment. Write down the decision and your reasoning when the emotion is high, then re-read it the next day before committing — you'll be shocked how often the second read changes your mind.",
     "Embedding Model Drift",
     "Every embedding model maps text into a vector space that's specific to that model's training — two different models will place the same sentence in completely different, incompatible coordinate systems, even if both models are 'good.' This means if you upgrade your embedding model, every vector you've already stored becomes meaningless when compared against new queries embedded with the new model; similarity scores become garbage silently, with no error thrown. The safe migration path is to re-embed your entire corpus with the new model into a separate shadow index, validate retrieval quality against a test set, and only then atomically switch traffic over — never let old-model and new-model vectors coexist in the same searchable collection, because the system will return confident-looking nonsense with no warning."),

    ("Make the first step embarrassingly small", "The reason big goals stall isn't lack of motivation, it's that the first visible step feels too large to start, so the brain avoids starting at all. Whatever the goal, shrink the very first action down until it feels almost too easy to skip — not 'write the report' but 'open the document and write one sentence.' Momentum, not motivation, is what actually gets things finished, and momentum only requires a start small enough that resistance can't stop it.",
     "Safe Backfills",
     "A backfill — reprocessing months of historical data through a new or fixed pipeline — is one of the riskiest operations in data engineering because it touches enormous volumes of data all at once. The safe pattern processes one bounded chunk at a time (one day or one partition), checkpoints progress to a control table after each chunk succeeds, and writes results to a staging location before atomically swapping them into production. This means if the backfill fails halfway through — and at scale, something eventually will — you resume from the checkpoint instead of starting over, and a bug in your logic only corrupts the staging area, never live production data. The rule of thumb: if you can't pause a backfill mid-flight and resume safely, it isn't ready to run against real data yet."),

    ("Ask what you'd tell a friend in your position", "When stuck in your own problem, you tend to catastrophize and see only the constraints; imagining the exact same situation happening to a close friend instantly makes the advice more obvious and more compassionate. Next time you're stuck on a hard call, physically ask out loud: 'if my friend told me this exact situation, what would I tell them?' The distance of imagining someone else almost always cuts through the fog faster than thinking about it directly as your own problem.",
     "Prompt Caching",
     "LLM providers can cache the computation of a prompt prefix so that if the same prefix appears in a later request, the model skips reprocessing it and only computes the new tokens — this can cut both latency and cost substantially on high-volume workloads. To benefit from this, structure your prompts so the static, unchanging part (system instructions, tool definitions, few-shot examples, retrieved reference documents) comes first, and the variable part specific to this exact request (the user's actual message) comes last. If you interleave static and dynamic content, or put the variable part first, the cache breaks on every request and you pay full price every time. This is a free performance win that costs nothing except reordering your prompt template correctly."),

    ("Write the anxious thought down, then answer it", "Anxious thoughts feel enormous and undefeatable while they loop silently in your head, but they shrink dramatically the moment you write them down as a specific claim you can evaluate. Take the worry, write it as a testable sentence ('I will lose this client if I miss this deadline'), then write the actual evidence for and against it underneath. Most anxious thoughts don't survive contact with a page and a pen — they were never as solid as they felt while spinning silently.",
     "Schema Evolution",
     "Data schemas change over time as businesses evolve, and the difference between a safe change and an outage is whether it's additive or breaking. Adding a new nullable column is safe — old readers simply ignore it and new readers can use it. Renaming a column, changing its type, or removing it is breaking, because every downstream consumer reading that field silently fails or gets wrong data, often without any error being thrown at all. The fix is treating your schema like a versioned API: enforce a schema registry that validates every write against a contract, require breaking changes to bump a major version with a clear deprecation window, and never let a producer ship a breaking change straight to production without downstream consumers explicitly opting in first."),

    ("Set a real finishing line, not just a starting one", "Open-ended goals like 'get better at X' never feel done, so they generate quiet, chronic guilt without ever producing the satisfaction of completion. Convert every open-ended goal into one with an explicit finish line and date: not 'learn Spanish' but 'hold a 10-minute conversation in Spanish by October 1st.' Finish lines let you actually finish something, celebrate it, and move on — instead of carrying every unfinished aspiration as background weight indefinitely.",
     "Evaluation Before You Scale",
     "It's tempting to tweak a model, a prompt, or a retrieval pipeline based on a handful of examples that 'look better,' but without a fixed evaluation set, you cannot actually tell if a change helped or just moved the failure cases around. Build a golden dataset of 50-200 real, labeled examples before you optimize anything — inputs paired with the correct or acceptable outputs, ideally pulled from real production failures. Score every candidate change against this exact same set with a fixed metric, and only ship changes that measurably improve the score, not ones that merely feel better on the three examples you happened to glance at. A model or prompt change without an eval harness isn't an improvement, it's a guess wearing the costume of one."),

    ("Lower the cost of saying no", "Every yes you give to a low-priority request is a no you're silently giving to your own priorities, but saying no feels costly because we imagine it will damage the relationship. In practice, a clear, kind, quick no ('I can't take this on right now, but here's who might be able to') costs far less goodwill than a slow, resentful yes followed by a late or half-hearted delivery. Practice a short, ready-made script for declining, and use it the moment a request doesn't align with what actually matters this week — the awkwardness fades in seconds; the time saved compounds for weeks.",
     "Late-Arriving Data",
     "Real-world events don't always arrive at your pipeline in the order they happened — a mobile app might batch and upload events hours after they occurred due to being offline, meaning your 'yesterday' partition is still incomplete when you first process it. Watermarking solves this by defining an explicit allowed lateness window (for example, 'accept events up to 48 hours late') and reprocessing the affected partitions on a rolling basis rather than assuming a partition is final the moment its calendar day ends. The practical implementation: never mark a partition as immutable until its lateness window has fully closed, and build your downstream aggregations to tolerate being recomputed for the last few days, not just the newest one — otherwise your 'final' daily numbers will quietly and repeatedly be wrong."),

    ("Track inputs weekly, not just outcomes monthly", "Outcomes (revenue, weight, streak percentage) are lagging indicators that only tell you the story of weeks ago; by the time an outcome moves, the inputs that caused it are long past being adjustable. Pick 2-3 leading inputs that you believe drive the outcome you care about, and track those weekly instead — hours of deep work, workouts completed, cold outreach sent. When the input trend is healthy but the outcome hasn't moved yet, that's normal lag, not failure, and it tells you to stay the course rather than panic and change strategy prematurely.",
     "Hybrid Search (Keyword + Vector)",
     "Pure vector search is excellent at matching meaning and paraphrases but can miss exact identifiers, product codes, or rare proper nouns that never appeared prominently in training data. Pure keyword search (BM25) is excellent at exact matches but misses semantically related content phrased differently. Hybrid search runs both retrievers in parallel and combines their ranked results using Reciprocal Rank Fusion (RRF) — a simple formula that rewards documents ranking highly in either list without needing to calibrate two different scoring scales against each other. In practice, adding RRF-based hybrid search on top of an existing vector-only system is often a two-line change that measurably improves retrieval quality, because the two methods fail in different, complementary places."),

    ("Design your environment instead of relying on willpower", "Willpower is a depletable resource that fails predictably under stress or fatigue, but your environment acts on you constantly without needing any willpower at all. If you want to stop checking your phone at night, don't rely on discipline — physically charge it in another room. If you want to eat better, don't rely on resisting temptation — don't keep the tempting food in the house at all. Redesigning the environment around the behavior you want is more reliable than any amount of trying harder, because it removes the decision point entirely.",
     "Orchestration & DAG Design",
     "Workflow orchestrators like Airflow or Dagster represent pipelines as directed acyclic graphs (DAGs) of tasks, and the single most important design rule is that every task must be parameterized by its logical execution date, never by reading the system clock with now() at runtime. A task that calls now() produces a different result every time it runs, which means it can never be safely replayed — and replay is exactly what you need when a bug is discovered three days later and you must reprocess that date's data. Keep individual tasks small and single-purpose rather than one giant script, since smaller tasks can be retried, monitored, and reasoned about independently, and a failure in one doesn't force you to redo unrelated work that already succeeded."),

    ("Ask for the specific help you need, not a vague ask", "Vague requests for help ('let me know if you have any thoughts') get vague, low-effort responses, because the other person doesn't know what would actually be useful to you. Instead, ask for the exact thing you need: 'can you review just the pricing section and tell me if the logic makes sense by Thursday?' Specific asks are both easier to say yes to and far more likely to get you something genuinely useful back, because you've done the thinking about what help actually looks like instead of outsourcing that thinking too.",
     "Data Contracts",
     "A data contract is an explicit, enforced agreement between the team producing data and the teams consuming it — covering the schema, the semantic meaning of each field, freshness SLAs, and who owns fixing it when something breaks. Without a contract, a producer team can rename a column or change its units to fix their own internal need, and every downstream dashboard or model silently breaks with no warning to anyone. Enforcing the contract in CI means the producer's build fails before merge if a breaking change isn't explicitly versioned and communicated — turning a silent 2am production incident into a visible, blocking conversation during code review, which is a strictly better time to have it."),

    ("Batch your reactive work instead of living in it", "Constantly checking messages and email throughout the day fragments your attention into pieces too small to do meaningful work in between, even though each individual check feels harmless. Instead, pick 2-3 fixed windows per day to process messages in a batch, and turn off notifications for the rest of the day entirely. The volume of reactive work you handle stays roughly the same, but the quality of everything else you do in between improves dramatically once it's no longer being constantly interrupted.",
     "Change Data Capture (CDC)",
     "Instead of repeatedly querying a database with SELECT * to check what's changed — which is slow, resource-intensive, and easy to get wrong — Change Data Capture reads the database's write-ahead log directly, streaming every insert, update, and delete as it happens in the exact order it occurred. This puts near-zero additional load on the source database, since you're tailing a log it already writes for its own durability rather than running extra queries against live tables. Tools like Debezium implement this pattern for popular databases, turning your operational database into a real-time event stream that downstream systems can consume incrementally instead of re-scanning the whole table on a schedule."),

    ("Let go of what you cannot verify", "A large fraction of daily worry is about things you cannot actually check or confirm — what someone privately thinks of you, whether a decision made months ago was the 'right' one, how a future event will unfold. When you notice you're spinning on something unverifiable, name it explicitly as unverifiable and consciously redirect attention to the nearest thing you can actually act on or check right now. This doesn't make the uncertainty disappear, but it stops you from spending finite attention on a question that has no available answer today.",
     "Cross-Encoder Reranking",
     "Bi-encoder retrieval (standard vector search) embeds the query and documents separately and compares them with a fast similarity calculation, which scales to millions of documents but loses some nuance because the query and document never actually interact during scoring. A cross-encoder takes the query and a candidate document together as a single input and produces a much more accurate relevance score, but it's far too slow to run against your entire corpus for every query. The standard pattern is a two-stage pipeline: retrieve 50-100 candidates cheaply with a bi-encoder, then rerank just those candidates with a cross-encoder to get the final, much more accurate top results — combining the speed of one method with the accuracy of the other."),

    ("Separate 'urgent' from 'important' honestly", "Urgent things demand attention immediately (a ringing phone, a Slack ping) while important things actually move your life or work forward (deep work, health, key relationships) — and urgent things will always win your attention by default unless you deliberately protect time for important-but-not-urgent work. At the start of each day, identify the one important-but-not-urgent task most likely to get crowded out, and schedule it before anything urgent has a chance to claim that slot. Most people's lives are entirely consumed by what's urgent; the ones who make real progress are the ones who protect space for what's merely important.",
     "Medallion Architecture (Bronze/Silver/Gold)",
     "This is a layered approach to organizing a data lake: the bronze layer holds raw, unmodified data exactly as it arrived from source systems, preserved for full traceability and reprocessing. The silver layer holds cleaned, validated, and conformed data — deduplicated, correctly typed, with business rules applied, but still close to the grain of the original events. The gold layer holds business-ready, aggregated marts built specifically for consumption by dashboards, reports, or downstream applications. The discipline that makes this work: no consumer is ever allowed to read directly from bronze, and no business logic is written below the gold layer boundary — each layer has one clear job, which makes debugging dramatically easier because you always know which layer is responsible for which kind of correctness."),

    ("Run small experiments instead of big bets", "Committing fully to one big untested plan feels decisive, but it means you only get one shot to be right, and being wrong is expensive both in time and in confidence. Instead, structure important changes as small, fast, reversible experiments — try the new habit for one week, ship the feature to 5% of users, pilot the new process on one team — before committing at scale. Small experiments let you be wrong cheaply and often, which paradoxically makes you right far more often over time than betting everything on being right the first time.",
     "Temperature & Top-p Sampling",
     "Temperature controls how much an LLM's output probabilities get rescaled before a token is sampled — near 0 makes the model almost always pick its single most likely next token (deterministic, focused, good for extraction and structured tasks), while higher values flatten the distribution and allow less likely tokens to be chosen (more varied, good for brainstorming and creative writing). Top-p (nucleus sampling) instead truncates the pool of candidate tokens to the smallest set whose cumulative probability exceeds p, cutting off the long unlikely tail regardless of temperature. The practical guidance: tune one of these knobs at a time and hold the other near its default, because moving both simultaneously makes it nearly impossible to reason about which change caused which behavior in your outputs."),

    ("Notice the story you're telling about the event", "The same event — a canceled meeting, a curt reply — can be interpreted as a personal slight or as someone simply having a busy day, and your emotional reaction follows the interpretation, not the event itself. When you feel a strong negative reaction, pause and ask: 'what's the story I've added on top of the bare fact, and what's at least one other plausible story?' This doesn't mean forcing false positivity — it means noticing that you chose an interpretation, often the most threatening one available, when a more neutral one fit the same facts just as well.",
     "Streaming vs Batch Processing",
     "True streaming processes each event within milliseconds to seconds of it occurring, but it demands much more operational complexity: stateful stream processors, exactly-once semantics, and infrastructure that must run continuously and be actively monitored around the clock. Micro-batch processing runs on a short fixed interval (every 1-5 minutes) and satisfies the vast majority of business requirements that get casually labeled 'real-time' at a fraction of the engineering and operational cost of true streaming. Before building a streaming system, ask explicitly what freshness the business actually requires in minutes, and only reach for true streaming when the answer is genuinely seconds — most of the time, the honest answer is minutes, and a simpler micro-batch system will serve it just as well."),

    ("Keep a record of decisions, not just outcomes", "When you only remember how something turned out, you tend to judge past decisions by results you couldn't have known in advance (hindsight bias), which teaches you the wrong lesson — a bad decision that got lucky looks good in hindsight, and a good decision that got unlucky looks bad. Keep a short log of significant decisions and your reasoning at the time you made them, before you know the outcome. Months later, review the log against what actually happened, and you'll learn to judge your decision-making process honestly, separate from luck.",
     "Structured Outputs from LLMs",
     "When an LLM's output feeds directly into another program rather than being read by a human, asking it nicely to 'please respond in JSON' is unreliable — it will occasionally add commentary, use inconsistent field names, or produce malformed JSON that breaks your parser. The reliable pattern is constraining the model's output directly: JSON schema enforcement or dedicated tool-calling modes offered by most providers guarantee the output conforms to a defined structure at the API level, not just via a polite request in the prompt. Even with structured output enforcement, validate the response against your schema in code before trusting it, and route validation failures to an automatic retry that includes the specific error message — this closes the loop and self-corrects most edge cases without any human intervention."),

    ("Give yourself permission to do it badly first", "Perfectionism at the start of a task is one of the most common causes of never starting at all, because the gap between 'blank page' and 'good' feels too large to cross in one step. Explicitly give yourself permission to produce a genuinely bad first draft — badly written, badly organized, wrong in places — with the sole goal of having something to react to and improve, rather than something to be proud of. Editing a bad draft into a good one is a completely different, far easier task than generating a good draft from nothing, and separating these two steps mentally is often the difference between finishing and stalling indefinitely.",
     "Dead Letter Queues",
     "In any pipeline processing large volumes of records, some small fraction will always be malformed, unexpected, or violate an assumption you didn't know you were making — and the question is not whether this happens but what happens when it does. A dead letter queue catches these failing records instead of crashing the whole pipeline or, worse, silently dropping them: the raw payload, the specific error encountered, and a timestamp all get written to a separate queue or table for later inspection. Alert on the DLQ's growth rate rather than its absolute size, since a sudden spike usually indicates a new upstream schema change or bug, while a small steady trickle might be an accepted, known edge case — the DLQ turns invisible data loss into a visible, monitorable, and replayable queue."),

    ("Match your effort to the actual stakes", "Not every decision deserves the same amount of deliberation — agonizing over a reversible, low-stakes choice (which of two similar restaurants to pick) wastes energy you need for the genuinely high-stakes, hard-to-reverse ones. Before deliberating, quickly classify the decision: if it's reversible and low-cost to undo, decide fast and move on; if it's hard to reverse and high-stakes, then it earns the careful analysis. Most decision fatigue comes from treating every choice as if it deserved the same weight as the rare, truly consequential ones.",
     "Pipeline Observability",
     "A pipeline can run 'successfully' — no errors thrown, job marked green — while still silently producing wrong or incomplete data, which is why success/failure status alone is not enough monitoring for a production data system. Track four signals on every important table: freshness (did it update within its expected window?), volume (is the row count within a normal historical band, not suddenly zero or doubled?), schema (has an unexpected column appeared or disappeared?), and distribution (are null rates and value ranges consistent with history?). Alert on deltas and anomalies relative to each table's own history, not on fixed absolute thresholds, since a table with naturally variable volume needs a different definition of 'normal' than one that's highly stable — this is what actually catches the 2am silent-corruption bugs that a green checkmark alone will never reveal."),

    ("Rest is part of the work, not a break from it", "Treating rest as merely the absence of work, something to feel guilty about, misunderstands how performance actually works — recovery is the phase where adaptation and consolidation actually happen, in training, in learning, and in creative work alike. Schedule real rest deliberately, the same way you schedule an important meeting, rather than letting it happen only as leftover time after everything else is done. A day of genuine, guilt-free rest taken proactively is usually more restorative and more sustainable than a burnt-out collapse taken reactively once you have no choice left.",
     "Context Window Budgeting",
     "An LLM's context window is a hard, finite resource, and treating it as unlimited by stuffing in entire conversation histories or full documents leads to degraded attention on the parts that actually matter, along with unnecessarily high token costs. Treat the context window like memory in an embedded system: allocate explicit, fixed budgets for the system prompt, retrieved reference material, and conversation history, and when the total would exceed the budget, summarize or evict the oldest turns rather than silently truncating from wherever the framework defaults to. This discipline both controls cost predictably and keeps the model's attention concentrated on the most relevant recent and retrieved content instead of diluted across everything it has ever seen in the conversation."),

    ("Plan for the version of you that won't feel like it", "The plan you make today, while motivated and clear-headed, will be executed later by a version of you that's tired, distracted, or simply doesn't feel like it — and plans that only work if motivation stays high are plans that will fail on exactly the days you need them most. Build systems and defaults that work even when willpower is at its lowest: lay out gym clothes the night before, pre-decide the workout so there's no decision left to make, automate the habit so 'starting' requires zero activation energy. The best plans assume low motivation as the default case, not the exception.",
     "Disaster Recovery for Data Systems",
     "Backups that have never actually been restored are only a theory of safety, not real safety, since the restore process itself often has its own undiscovered bugs that only surface when you desperately need it to work. Define two concrete numbers before an incident ever happens: your Recovery Point Objective (how much data loss is acceptable, e.g., the last hour) and your Recovery Time Objective (how long you can be down, e.g., 30 minutes) — then build and test a restore process against those actual targets. Table-format time-travel features (like those in Delta Lake or Iceberg) are excellent for undoing an accidental bad write, but they are not a substitute for real disaster recovery against a lost or corrupted underlying storage system — rehearse a full restore on a schedule, not just when you're desperate."),
]


# ---------------------------------------------------------------------------
# Excel persistence
# ---------------------------------------------------------------------------
def load_log(path: str) -> pd.DataFrame:
    if os.path.exists(path):
        try:
            df = pd.read_excel(path, dtype={"Date": str})
            for col in COLUMNS:
                if col not in df.columns:
                    df[col] = None
            return df[COLUMNS]
        except Exception as e:
            st.error(f"Could not read existing log ({e}). A fresh one will be created on save.")
    return pd.DataFrame(columns=COLUMNS)


def compute_streak(df: pd.DataFrame, today_key: str, today_xp: int) -> tuple[int, int]:
    hist = df[df["Date"] != today_key].copy()
    hist = hist.sort_values("Date")
    xp_by_date = dict(zip(hist["Date"], hist["Daily XP %"]))
    xp_by_date[today_key] = today_xp

    dates_sorted = sorted(xp_by_date.keys())
    streaks, run = [], 0
    for d in dates_sorted:
        if xp_by_date.get(d, 0) >= 100:
            run += 1
        else:
            if run:
                streaks.append(run)
            run = 0
    if run:
        streaks.append(run)

    current = 0
    for d in reversed(dates_sorted):
        if xp_by_date.get(d, 0) >= 100:
            current += 1
        else:
            break
    best = max(streaks) if streaks else 0
    best = max(best, current)
    return current, best


def write_workbook(path: str, df: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "Forge Log"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="20242E", end_color="20242E")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for _, r in df.iterrows():
        ws.append([r[c] for c in COLUMNS])

    widths = [12, 11, 20, 13, 22, 20, 24, 18, 13, 18, 34, 11, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(path)


def save_entry(path: str, row: dict):
    df = load_log(path)
    today_key = row["Date"]
    df = df[df["Date"] != today_key]
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    df = df.sort_values("Date").reset_index(drop=True)
    write_workbook(path, df)
    return df


def delete_today(path: str, today_key: str):
    df = load_log(path)
    df = df[df["Date"] != today_key].reset_index(drop=True)
    write_workbook(path, df)
    return df


# ---------------------------------------------------------------------------
# Style — dark minimalist, violet accent, activity ring, micro-interactions
# ---------------------------------------------------------------------------
def inject_style():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@600;700;800&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@400;500&display=swap');

        :root{
            --ink:#0a0a0d; --card:#151519; --card-hi:#1c1c22;
            --text:#f5f5f7; --muted:#93939e; --line:rgba(255,255,255,.08);
            --accent:#7c6cf0; --accent-2:#c26cf0; --accent-soft:rgba(124,108,240,.14);
            --ok:#34d399; --ok-soft:rgba(52,211,153,.14);
        }

        html, body, [class*="css"] { font-family:'Inter',sans-serif; }
        .stApp { background:var(--ink); color:var(--text); }
        h1,h2,h3,h4 { font-family:'Manrope',sans-serif !important; letter-spacing:-.01em; }

        #MainMenu, footer, header[data-testid="stHeader"] { background:transparent; }
        .block-container { padding-top:1.4rem; padding-bottom:3rem; max-width:640px; }

        .hero { text-align:center; margin-bottom:1rem; animation:fadeDown .5s ease; }
        .hero .mark { font-size:2.2rem; font-weight:800; letter-spacing:-.02em; line-height:1; }
        .hero .mark span{ color:var(--accent); }
        .hero .tag { font-family:'JetBrains Mono',monospace; font-size:.65rem; letter-spacing:.3em;
                     color:var(--muted); text-transform:uppercase; margin-top:.3rem; }
        @keyframes fadeDown{ from{opacity:0; transform:translateY(-8px);} to{opacity:1; transform:none;} }

        div[data-testid="stVerticalBlockBorderWrapper"]{
            background:var(--card) !important; border:1px solid var(--line) !important;
            border-radius:16px !important; transition:transform .18s ease, border-color .18s ease, box-shadow .18s ease;
            animation:rise .45s ease backwards;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:hover{
            border-color:rgba(124,108,240,.35) !important; box-shadow:0 6px 24px rgba(0,0,0,.28);
        }
        @keyframes rise{ from{opacity:0; transform:translateY(10px);} to{opacity:1; transform:none;} }

        div[data-testid="stTabs"] button[role="tab"]{
            font-family:'Manrope',sans-serif; font-weight:700; font-size:.85rem;
            border-radius:999px !important; padding:.4rem 1rem !important; transition:all .18s ease; color:var(--muted);
        }
        div[data-testid="stTabs"] button[role="tab"]:hover{ color:var(--text); background:var(--card-hi); }
        div[data-testid="stTabs"] button[aria-selected="true"]{ color:#fff !important; background:var(--accent) !important; }
        div[data-testid="stTabs"] [data-baseweb="tab-highlight"]{ display:none; }
        div[data-testid="stTabs"] [data-baseweb="tab-border"]{ background:transparent; }

        label[data-testid="stCheckbox"]{ transition:transform .12s ease; }
        label[data-testid="stCheckbox"]:hover{ transform:translateX(2px); }

        div[data-testid="stButton"] button{
            border-radius:11px !important; font-weight:700 !important;
            transition:transform .1s ease, box-shadow .18s ease !important; border:1px solid var(--line) !important;
        }
        div[data-testid="stButton"] button:hover{ transform:translateY(-1px); box-shadow:0 6px 18px rgba(124,108,240,.18); }
        div[data-testid="stButton"] button:active{ transform:translateY(0) scale(.98); }

        div[data-testid="stMetric"]{
            background:var(--card-hi); border:1px solid var(--line); border-radius:12px;
            padding:.7rem .9rem; transition:transform .15s ease;
        }
        div[data-testid="stMetric"]:hover{ transform:translateY(-2px); }
        div[data-testid="stMetricLabel"]{ font-family:'JetBrains Mono',monospace; text-transform:uppercase;
            font-size:.64rem !important; letter-spacing:.14em; color:var(--muted) !important; }

        div[data-testid="stNumberInput"] input, div[data-testid="stTextArea"] textarea{
            border-radius:10px !important; background:var(--card-hi) !important; border:1px solid var(--line) !important;
        }

        section[data-testid="stSidebar"]{ background:var(--card) !important; border-right:1px solid var(--line); }
        hr{ border-color:var(--line) !important; margin:1.1rem 0 !important; }

        .quote-card{ font-family:'Manrope',sans-serif; font-size:1.05rem; font-weight:700; line-height:1.5; padding:.2rem 0 .1rem; }
        .quote-card .q::before{ content:"“"; color:var(--accent); font-size:1.4rem; }
        .quote-author{ font-family:'JetBrains Mono',monospace; font-size:.76rem; color:var(--muted); margin-top:.2rem; }

        .tier-chip{
            display:inline-block; font-family:'JetBrains Mono',monospace; font-size:.7rem; font-weight:700;
            letter-spacing:.1em; text-transform:uppercase; padding:.3rem .7rem; border-radius:999px;
            background:var(--accent-soft); color:var(--accent); border:1px solid rgba(124,108,240,.3);
        }
        .tier-chip.sage{ background:var(--ok-soft); color:var(--ok); border-color:rgba(52,211,153,.3); }

        .sync-badge{
            display:flex; align-items:center; gap:7px; font-family:'JetBrains Mono',monospace; font-size:.72rem;
            color:var(--ok); background:var(--ok-soft); border:1px solid rgba(52,211,153,.3);
            border-radius:10px; padding:.5rem .7rem; margin-top:.6rem;
        }
        .sync-dot{ width:7px; height:7px; border-radius:50%; background:var(--ok); flex:none;
            animation:pulse 1.6s ease-in-out infinite; }
        @keyframes pulse{ 0%,100%{opacity:1} 50%{opacity:.4} }
        </style>
        """,
        unsafe_allow_html=True,
    )


def ring_svg(pct: int, size: int = 88) -> str:
    r = 38
    circumference = 2 * 3.14159265 * r
    offset = circumference - (pct / 100) * circumference
    return f"""
    <div style="position:relative;width:{size}px;height:{size}px;flex:none;">
      <svg viewBox="0 0 88 88" width="{size}" height="{size}" style="transform:rotate(-90deg);">
        <defs>
          <linearGradient id="ringGrad" x1="0%" y1="0%" x2="100%" y2="100%">
            <stop offset="0%" stop-color="#7c6cf0"/>
            <stop offset="100%" stop-color="#c26cf0"/>
          </linearGradient>
        </defs>
        <circle cx="44" cy="44" r="{r}" fill="none" stroke="#1c1c22" stroke-width="9"/>
        <circle cx="44" cy="44" r="{r}" fill="none" stroke="url(#ringGrad)" stroke-width="9"
                stroke-linecap="round" stroke-dasharray="{circumference:.2f}" stroke-dashoffset="{offset:.2f}"
                style="transition:stroke-dashoffset .6s cubic-bezier(.22,1,.36,1);"/>
      </svg>
      <div style="position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;">
        <div style="font-family:'Manrope',sans-serif;font-weight:800;font-size:19px;color:#f5f5f7;">{pct}%</div>
        <div style="font-size:9px;color:#5c5c66;letter-spacing:.08em;text-transform:uppercase;margin-top:-2px;">XP</div>
      </div>
    </div>
    """


# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
inject_style()

st.markdown(
    """
    <div class="hero">
        <div class="mark">F<span>O</span>RGE</div>
        <div class="tag">The Daily Ledger</div>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.markdown("#### ⚙ Settings")
    log_path = st.text_input("Excel log file path", value=DEFAULT_LOG_PATH, label_visibility="collapsed")
    st.caption("Every change auto-saves here — no button needed.")
    if os.path.exists(log_path):
        st.success(f"Log found · {os.path.getsize(log_path)/1024:.1f} KB", icon="📗")
    else:
        st.info("No log yet — created on your first change.", icon="📄")

today = dt.date.today()
today_key = today.isoformat()
py_wd = today.weekday()
split_title, split_detail = SPLIT_BY_PY_WEEKDAY[py_wd]
day_of_year = today.timetuple().tm_yday
life_title, life_body, lesson_title, lesson_body = INTEL[day_of_year % len(INTEL)]

existing_df = load_log(log_path)

if st.session_state.get("_clear_today"):
    st.session_state["_clear_today"] = False
    delete_today(log_path, today_key)
    st.session_state["reset_gen"] = st.session_state.get("reset_gen", 0) + 1
    existing_df = load_log(log_path)

reset_gen = st.session_state.get("reset_gen", 0)
wkey = lambda name: f"{name}_{today_key}_{reset_gen}"

existing_today = existing_df[existing_df["Date"] == today_key]
prefill = existing_today.iloc[0].to_dict() if not existing_today.empty else {}

st.caption(f"📅 {today.strftime('%A, %B %d, %Y')}")

tab_today, tab_journal, tab_intel, tab_stats = st.tabs(["🏠 Today", "✍ Journal", "🧠 Intel", "📊 Stats"])

with tab_today:
    with st.container(border=True):
        st.markdown("##### ⚒ Today's Iron")
        st.markdown(f"**{split_title}**")
        st.caption(split_detail)
        workout_done = st.checkbox("Mark today's session complete", value=bool(prefill.get("Workout Done", False)), key=wkey("workout_done"))

    with st.container(border=True):
        st.markdown("##### 🎯 Rule of Three")
        t1 = st.checkbox("Code / review pipeline architecture layouts", value=bool(prefill.get("Target 1 - Code/Review", False)), key=wkey("t1"))
        t2 = st.checkbox("Deep work window (no distractions)", value=bool(prefill.get("Target 2 - Deep Work", False)), key=wkey("t2"))
        t3 = st.checkbox("Drink 4L of water & clean nutrition", value=bool(prefill.get("Target 3 - Water/Nutrition", False)), key=wkey("t3"))

    with st.container(border=True):
        st.markdown("##### ✦ Nightly Reflection")
        gratitude = st.checkbox("I've thought of 3 things I'm grateful for", value=bool(prefill.get("Gratitude Reflected", False)), key=wkey("gratitude"))

with tab_journal:
    with st.container(border=True):
        st.markdown("##### ⏱ Time Ledger")
        col1, col2 = st.columns(2)
        with col1:
            project_hours = st.number_input(
                "Project hours", min_value=0.0, max_value=24.0, step=0.25,
                value=float(prefill.get("Project Hours", 0.0) or 0.0), key=wkey("project_hours"),
            )
        with col2:
            study_hours = st.number_input(
                "Study / review hours", min_value=0.0, max_value=24.0, step=0.25,
                value=float(prefill.get("Study/Review Hours", 0.0) or 0.0), key=wkey("study_hours"),
            )

    with st.container(border=True):
        st.markdown("##### 📝 What did you do today?")
        journal = st.text_area(
            "Journal", value=str(prefill.get("Journal", "") or ""), height=110,
            placeholder="Shipped the ingestion retry logic, reviewed two PRs, read a chapter on vector indexing…",
            label_visibility="collapsed", max_chars=600, key=wkey("journal"),
        )
        st.caption(f"{len(journal)} / 600")

with tab_intel:
    with st.container(border=True):
        st.markdown("##### 🌱 Today's Practice")
        st.markdown(f'<div class="quote-card">{life_title}</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="quote-author" style="line-height:1.6;">{life_body}</div>', unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown("##### 🧠 Today's Lesson")
        st.markdown(f"**{lesson_title}**")
        st.markdown(f'<div style="font-size:.92rem; line-height:1.65; color:var(--text);">{lesson_body}</div>', unsafe_allow_html=True)

with tab_stats:
    df_display = load_log(log_path)
    if df_display.empty:
        st.info("No entries yet — your first save will appear here automatically.")
    else:
        with st.container(border=True):
            total_project = pd.to_numeric(df_display["Project Hours"], errors="coerce").sum()
            total_study = pd.to_numeric(df_display["Study/Review Hours"], errors="coerce").sum()
            cc1, cc2, cc3 = st.columns(3)
            cc1.metric("Days logged", len(df_display))
            cc2.metric("Project hrs", f"{total_project:.1f}")
            cc3.metric("Study hrs", f"{total_study:.1f}")
        st.dataframe(df_display.sort_values("Date", ascending=False), width='stretch', hide_index=True)

    with st.container(border=True):
        st.markdown("##### 🗓 Day Cycle")
        st.caption(
            "Each calendar day gets its own row automatically — tomorrow starts blank "
            "with no action needed. Use this only if you want to wipe *today's* entry and start over."
        )
        if st.button("🗑 Clear today's entry", width='stretch'):
            st.session_state["_clear_today"] = True
            st.rerun()

# ---- XP + streak ----
tokens = [workout_done, t1, t2, t3, gratitude]
xp = sum(tokens) * 20
tier_label = "Sovereign Sage" if xp >= 100 else ("Warrior-Scholar" if xp >= 40 else "The Seeker")
tier_class = "sage" if xp >= 100 else ""

current_streak, best_streak = compute_streak(existing_df, today_key, xp)

with st.sidebar:
    st.markdown("---")
    st.markdown(ring_svg(xp), unsafe_allow_html=True)
    st.markdown(f'<div style="text-align:center;margin-top:8px;"><span class="tier-chip {tier_class}">{tier_label}</span></div>', unsafe_allow_html=True)
    s1, s2 = st.columns(2)
    s1.metric("Streak", f"{current_streak}d")
    s2.metric("Best", f"{best_streak}d")

# ---- Auto-save on every rerun (no button — feels like a live database) ----
row = {
    "Date": today_key,
    "Weekday": today.strftime("%A"),
    "Workout Split": split_title,
    "Workout Done": workout_done,
    "Target 1 - Code/Review": t1,
    "Target 2 - Deep Work": t2,
    "Target 3 - Water/Nutrition": t3,
    "Gratitude Reflected": gratitude,
    "Project Hours": project_hours,
    "Study/Review Hours": study_hours,
    "Journal": journal,
    "Daily XP %": xp,
    "Streak": current_streak,
}

try:
    save_entry(log_path, row)
    with st.sidebar:
        st.markdown(
            f'<div class="sync-badge"><span class="sync-dot"></span>Synced to Excel · {dt.datetime.now().strftime("%H:%M:%S")}</div>',
            unsafe_allow_html=True,
        )
    if xp >= 100 and not st.session_state.get("_celebrated_today") == today_key:
        st.session_state["_celebrated_today"] = today_key
        st.balloons()
        st.toast("Perfect day — 100% XP!", icon="🔥")
except Exception as e:
    with st.sidebar:
        st.error(f"Auto-save failed: {e}")

if st.session_state.get("_force_reset"):
    st.session_state["_force_reset"] = False
    st.toast("Day committed. Streak updates on your next perfect day.", icon="✅")
